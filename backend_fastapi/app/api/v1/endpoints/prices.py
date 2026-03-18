from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.entities import AuditLog, Marketplace, MarketplaceStore, Price, Product, User
from app.services.marketplaces import OzonClient
from app.services.sync import get_store_credentials

from app.schemas.price import (
    PriceListOut,
    PriceLogEntryOut,
    PricePatchIn,
    PriceRowOut,
    PricesApplyMarkupIn,
    PricesBulkUpdateIn,
    PricesReloadIn,
    PricesReloadOut,
)

router = APIRouter(prefix='/prices', tags=['prices'])

ACQUIRING_RATE = 0.019
PACKAGING_COST = 40.0
PROMOTION_RATE = 0.01
DEFAULT_COMMISSION_PERCENT = 12.0
DEFAULT_CUSTOMER_DELIVERY = 30.0
DEFAULT_LOGISTICS = 55.0
DEFAULT_FIRST_MILE = 12.0
DEFAULT_FBS_COST = 15.0


def _calculate_price_metrics(
    current_price: float,
    cost_price: float,
    customer_delivery: float,
    logistics: float,
    first_mile: float,
    packaging: float,
    promotion_percent: float,
    ozon_commission_percent: float,
    fbs_cost: float,
) -> dict[str, float]:
    current_price = round(max(current_price, 0.0), 2)
    cost_price = round(max(cost_price, 0.0), 2)
    promotion_percent = round(max(promotion_percent, 0.0), 4)

    acquiring = round(current_price * ACQUIRING_RATE, 2)
    promotion = round(current_price * promotion_percent / 100, 2)
    ozon_commission_rub = round(current_price * ozon_commission_percent / 100, 2)
    payout_to_seller = round(
        current_price
        - acquiring
        - customer_delivery
        - logistics
        - first_mile
        - packaging
        - promotion
        - ozon_commission_rub
        - fbs_cost,
        2,
    )
    margin_rub = round(payout_to_seller - cost_price, 2)
    margin_percent = round((margin_rub / current_price) * 100, 2) if current_price else 0.0
    markup_percent = round(((payout_to_seller / cost_price) - 1) * 100, 2) if cost_price else 0.0

    return {
        'acquiring': acquiring,
        'promotion': promotion,
        'ozon_commission_rub': ozon_commission_rub,
        'payout_to_seller': payout_to_seller,
        'margin_rub': margin_rub,
        'margin_percent': margin_percent,
        'markup_percent': markup_percent,
    }


def _calculate_price_from_markup(
    markup_percent: float,
    cost_price: float,
    customer_delivery: float,
    logistics: float,
    first_mile: float,
    packaging: float,
    promotion_percent: float,
    ozon_commission_percent: float,
    fbs_cost: float,
) -> float:
    denominator = 1 - ACQUIRING_RATE - (promotion_percent / 100) - (ozon_commission_percent / 100)
    if denominator <= 0:
        return 0.0
    payout_target = cost_price * (1 + markup_percent / 100)
    return round((payout_target + customer_delivery + logistics + first_mile + packaging + fbs_cost) / denominator, 2)


def _build_price_row(product: Product, price: Price, stock: int = 0, fbs: int = 0, fbo: int = 0) -> PriceRowOut:
    current_price = float(price.current_price)
    min_price = round(max(_to_float(price.min_price, current_price), 0.0), 2)
    ozon_data = price.ozon_data or {}
    price_payload = ozon_data.get('price') or {}
    commissions = ozon_data.get('commissions') or {}

    marketing_seller_price = round(max(current_price, 0.0), 2)
    cost_price = _to_float(price.cost_price)
    if cost_price <= 0:
        fallback_cost = _to_float(price_payload.get('net_price'), float(price.previous_price) if price.previous_price else 0.0)
        cost_price = fallback_cost if fallback_cost > 0 else round(marketing_seller_price * 0.7, 2)

    customer_delivery = round(_to_float(commissions.get('fbs_deliv_to_customer_amount'), DEFAULT_CUSTOMER_DELIVERY), 2)
    logistics = round(_to_float(commissions.get('fbs_direct_flow_trans_min_amount'), DEFAULT_LOGISTICS), 2)
    ozon_commission_percent = round(_to_float(commissions.get('sales_percent_fbs'), DEFAULT_COMMISSION_PERCENT), 2)
    first_mile = round(_to_float(ozon_data.get('first_mile'), DEFAULT_FIRST_MILE), 2)
    packaging = round(_to_float(ozon_data.get('packaging'), PACKAGING_COST), 2)
    promotion_percent = round(_to_float(ozon_data.get('promotion_percent'), PROMOTION_RATE * 100), 2)
    fbs_cost = round(_to_float(ozon_data.get('fbs_cost'), DEFAULT_FBS_COST), 2)

    metrics = _calculate_price_metrics(
        marketing_seller_price,
        cost_price,
        customer_delivery,
        logistics,
        first_mile,
        packaging,
        promotion_percent,
        ozon_commission_percent,
        fbs_cost,
    )

    return PriceRowOut(
        product_id=product.id,
        offer_id=product.article or product.sku,
        title=product.title,
        stock=stock,
        fbs=fbs,
        fbo=fbo,
        current_price=marketing_seller_price,
        min_price=min_price,
        previous_price=price.previous_price,
        acquiring=metrics['acquiring'],
        customer_delivery=customer_delivery,
        logistics=logistics,
        first_mile=first_mile,
        packaging=packaging,
        promotion=metrics['promotion'],
        promotion_percent=promotion_percent,
        ozon_commission_percent=ozon_commission_percent,
        ozon_commission_rub=metrics['ozon_commission_rub'],
        cost_price=cost_price,
        fbs_cost=fbs_cost,
        payout_to_seller=metrics['payout_to_seller'],
        markup_percent=metrics['markup_percent'],
        margin_rub=metrics['margin_rub'],
        margin_percent=metrics['margin_percent'],
    )


def _get_owned_ozon_store_ids(db: Session, user_id: int) -> list[int]:
    return db.scalars(
        select(MarketplaceStore.id).where(
            MarketplaceStore.user_id == user_id,
            MarketplaceStore.is_enabled.is_(True),
            MarketplaceStore.marketplace == Marketplace.ozon,
        )
    ).all()


def _log_price_action(db: Session, user_id: int, action: str, details: dict):
    db.add(
        AuditLog(
            user_id=user_id,
            action=action,
            entity_type='prices',
            entity_id=str(details.get('store_id', 'global')),
            details=details,
        )
    )


def _to_float(value, default: float = 0.0) -> float:
    try:
        if value is None or value == '':
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_str(value, default: str = '') -> str:
    if value is None:
        return default
    return str(value).strip() or default


def _upsert_products_and_prices(
    db: Session,
    store_id: int,
    price_items: list[dict],
    now: datetime,
) -> tuple[int, int, int]:
    offer_ids = list(dict.fromkeys(_to_str(item.get('offer_id')) for item in price_items if _to_str(item.get('offer_id'))))
    if not offer_ids:
        return 0, 0, 0

    existing_products = db.scalars(
        select(Product).where(Product.store_id == store_id, or_(Product.article.in_(offer_ids), Product.sku.in_(offer_ids)))
    ).all()
    products_by_offer: dict[str, Product] = {}
    for product in existing_products:
        if product.article:
            products_by_offer[product.article] = product
        if product.sku:
            products_by_offer[product.sku] = product

    created_products = 0
    touched_product_ids: set[int] = set()

    for item in price_items:
        offer_id = _to_str(item.get('offer_id'))
        if not offer_id:
            continue
        product_id = _to_str(item.get('product_id'))
        title = _to_str(item.get('name'), default='Ozon product')

        product = products_by_offer.get(offer_id)
        if product is None:
            product = Product(store_id=store_id, sku=offer_id, article=offer_id, title=title)
            db.add(product)
            db.flush()
            products_by_offer[offer_id] = product
            created_products += 1
        else:
            if title and product.title != title:
                product.title = title
            if product_id and product.sku != product_id:
                product.sku = product_id
            if not product.article:
                product.article = offer_id

        touched_product_ids.add(product.id)

    existing_prices = db.scalars(select(Price).where(Price.product_id.in_(touched_product_ids))).all()
    prices_by_product_id = {price.product_id: price for price in existing_prices}

    created_prices = 0
    for item in price_items:
        offer_id = _to_str(item.get('offer_id'))
        if not offer_id:
            continue
        product = products_by_offer.get(offer_id)
        if product is None:
            continue

        price_payload = item.get('price') or {}
        current_price = _to_float(price_payload.get('marketing_seller_price'), _to_float(price_payload.get('price')))
        min_price = _to_float(price_payload.get('min_price'), current_price)
        previous_price = _to_float(price_payload.get('old_price'), None)

        price = prices_by_product_id.get(product.id)
        if price is None:
            cost_price = _to_float(price_payload.get('net_price'))
            if cost_price <= 0:
                cost_price = round(current_price * 0.7, 2)
            db.add(
                Price(
                    product_id=product.id,
                    current_price=current_price,
                    min_price=min_price,
                    previous_price=previous_price,
                    cost_price=cost_price,
                    ozon_data=item,
                    updated_at=now,
                )
            )
            created_prices += 1
            continue

        price.previous_price = previous_price if previous_price is not None else price.current_price
        price.current_price = current_price
        price.min_price = min_price
        if _to_float(price.cost_price) <= 0:
            fallback_cost = _to_float(price_payload.get('net_price'))
            if fallback_cost > 0:
                price.cost_price = round(fallback_cost, 2)
        price.ozon_data = item
        price.updated_at = now

    return len(touched_product_ids), created_products, created_prices


def _fetch_logs(db: Session, user_id: int, limit: int = 20) -> list[PriceLogEntryOut]:
    logs = db.scalars(
        select(AuditLog)
        .where(AuditLog.user_id == user_id, AuditLog.entity_type == 'prices')
        .order_by(AuditLog.created_at.desc())
        .limit(limit)
    ).all()
    return [
        PriceLogEntryOut(
            id=log.id,
            message=f"{log.action}: {log.details}",
            created_at=log.created_at.isoformat(),
        )
        for log in logs
    ]


def _prices_query(db: Session, store_id: int, search: str | None):
    q = (
        select(Product, Price)
        .join(Price, Price.product_id == Product.id)
        .where(Product.store_id == store_id)
    )

    if search:
        needle = f'%{search.strip()}%'
        q = q.where(or_(Product.sku.ilike(needle), Product.article.ilike(needle), Product.title.ilike(needle)))

    return q


@router.get('', response_model=PriceListOut)
def list_prices(
    store_id: int = Query(...),
    search: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=100, ge=1, le=5000),
    sort_by: str = Query(default='stock'),
    sort_order: str = Query(default='desc'),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    owned_store_ids = _get_owned_ozon_store_ids(db, current_user.id)
    if store_id not in owned_store_ids:
        raise HTTPException(status_code=404, detail='Store not found')

    q = _prices_query(db, store_id, search)
    total = db.scalar(select(func.count()).select_from(q.subquery())) or 0

    rows = db.execute(q).all()

    offer_ids = [product.article or product.sku for product, _price in rows]
    credentials = get_store_credentials(db, store_id)
    stocks_map: dict[str, dict[str, int]] = {}
    try:
        stocks_map = OzonClient().fetch_stocks(credentials, offer_ids)
    except Exception:
        stocks_map = {}

    items: list[PriceRowOut] = []
    for product, price in rows:
        offer = product.article or product.sku
        stock_data = stocks_map.get(offer, {'fbs': 0, 'fbo': 0})
        fbs = int(stock_data.get('fbs', 0))
        fbo = int(stock_data.get('fbo', 0))
        items.append(_build_price_row(product, price, fbs + fbo, fbs, fbo))

    if sort_by == 'price':
        items.sort(key=lambda i: i.current_price, reverse=sort_order == 'desc')
    elif sort_by == 'offer_id':
        items.sort(key=lambda i: i.offer_id, reverse=sort_order == 'desc')
    else:
        items.sort(key=lambda i: i.stock, reverse=sort_order == 'desc')

    paged_items = items[(page - 1) * page_size : (page - 1) * page_size + page_size]

    logs = _fetch_logs(db, current_user.id)
    return PriceListOut(total=total, page=page, page_size=page_size, items=paged_items, logs=logs)


@router.post('/reload', response_model=PricesReloadOut)
def reload_prices(
    payload: PricesReloadIn,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    owned_store_ids = _get_owned_ozon_store_ids(db, current_user.id)
    if payload.store_id not in owned_store_ids:
        raise HTTPException(status_code=404, detail='Store not found')

    store = db.scalar(select(MarketplaceStore).where(MarketplaceStore.id == payload.store_id))
    if store is None or store.marketplace != Marketplace.ozon:
        raise HTTPException(status_code=404, detail='Store not found')

    credentials = get_store_credentials(db, payload.store_id)
    client = OzonClient()

    try:
        offer_ids = client.fetch_all_offer_ids(credentials)
        price_items = client.fetch_prices(credentials, offer_ids)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f'Failed to load prices from Ozon API: {exc}') from exc

    now = datetime.utcnow()
    touched_items, created_products, created_prices = _upsert_products_and_prices(db, payload.store_id, price_items, now)

    _log_price_action(
        db,
        current_user.id,
        'reload',
        {
            'store_id': payload.store_id,
            'items': touched_items,
            'created_products': created_products,
            'created_prices': created_prices,
            'fetched_offer_ids': len(offer_ids),
            'fetched_price_items': len(price_items),
            'at': now.isoformat(),
        },
    )
    db.commit()
    return PricesReloadOut(status='ok', message=f'Загружено товаров: {touched_items}, цен: {len(price_items)}')


@router.patch('/{offer_id}', response_model=PriceRowOut)
def update_price(
    offer_id: str,
    payload: PricePatchIn,
    store_id: int = Query(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    owned_store_ids = _get_owned_ozon_store_ids(db, current_user.id)
    if store_id not in owned_store_ids:
        raise HTTPException(status_code=404, detail='Store not found')

    row = db.execute(
        select(Product, Price)
        .join(Price, Price.product_id == Product.id)
        .where(
            Product.store_id == store_id,
            or_(Product.article == offer_id, Product.sku == offer_id),
        )
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail='Offer not found')

    product, price = row

    store_credentials = get_store_credentials(db, store_id)
    client = OzonClient()

    ozon_data = price.ozon_data or {}
    price_payload = ozon_data.get('price') or {}
    commissions = ozon_data.get('commissions') or {}
    cost_price = payload.cost_price if payload.cost_price is not None else _to_float(
        price.cost_price,
        _to_float(price_payload.get('net_price'), float(price.previous_price) if price.previous_price else float(price.current_price) * 0.7),
    )
    packaging = payload.packaging if payload.packaging is not None else _to_float(ozon_data.get('packaging'), PACKAGING_COST)
    promotion_percent = payload.promotion_percent if payload.promotion_percent is not None else _to_float(
        ozon_data.get('promotion_percent'),
        PROMOTION_RATE * 100,
    )

    target_price = payload.new_price
    target_min_price = round(max(payload.min_price if payload.min_price is not None else _to_float(price.min_price, payload.new_price), 1.0), 2)
    if payload.markup_percent is not None:
        target_price = _calculate_price_from_markup(
            payload.markup_percent,
            cost_price,
            _to_float(commissions.get('fbs_deliv_to_customer_amount'), DEFAULT_CUSTOMER_DELIVERY),
            _to_float(commissions.get('fbs_direct_flow_trans_min_amount'), DEFAULT_LOGISTICS),
            _to_float(ozon_data.get('first_mile'), DEFAULT_FIRST_MILE),
            packaging,
            promotion_percent,
            _to_float(commissions.get('sales_percent_fbs'), DEFAULT_COMMISSION_PERCENT),
            _to_float(ozon_data.get('fbs_cost'), DEFAULT_FBS_COST),
        )

    target_price = round(max(target_price, 1.0), 2)
    target_min_price = round(min(target_price, max(target_min_price, 1.0)), 2)

    try:
        client.import_prices(
            store_credentials,
            [
                {
                    'offer_id': product.article or product.sku,
                    'product_id': int(_to_float(ozon_data.get('product_id'), 0)) or None,
                    'price': f'{target_price:.2f}',
                    'min_price': f'{target_min_price:.2f}',
                    'currency_code': _to_str(price_payload.get('currency_code'), 'RUB'),
                    'vat': _to_str(price_payload.get('vat'), '0.22'),
                }
            ],
        )
        refreshed = client.fetch_prices(store_credentials, [product.article or product.sku])
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail='Не удалось обновить цену в Ozon') from exc

    fresh_item = refreshed[0] if refreshed else {'price': {'marketing_seller_price': target_price}}
    fresh_price_payload = fresh_item.get('price') or {}
    updated_price = _to_float(fresh_price_payload.get('marketing_seller_price'), target_price)
    updated_min_price = _to_float(fresh_price_payload.get('min_price'), target_min_price)

    price.previous_price = price.current_price
    price.current_price = updated_price
    price.min_price = round(min(updated_price, max(updated_min_price, 1.0)), 2)
    if fresh_item:
        fresh_item['promotion_percent'] = promotion_percent
        fresh_item['packaging'] = packaging
        fresh_item['first_mile'] = _to_float(ozon_data.get('first_mile'), DEFAULT_FIRST_MILE)
        fresh_item['fbs_cost'] = _to_float(ozon_data.get('fbs_cost'), DEFAULT_FBS_COST)
    price.cost_price = round(max(cost_price, 0.0), 2)
    price.ozon_data = fresh_item or price.ozon_data
    price.updated_at = datetime.utcnow()

    _log_price_action(
        db,
        current_user.id,
        'patch',
        {
            'store_id': store_id,
            'offer_id': offer_id,
            'new_price': target_price,
            'min_price': target_min_price,
            'status': 'updated_in_ozon',
        },
    )

    db.commit()
    db.refresh(price)

    try:
        stocks_map = client.fetch_stocks(store_credentials, [product.article or product.sku])
        stock_data = stocks_map.get(product.article or product.sku, {'fbs': 0, 'fbo': 0})
        fbs = int(stock_data.get('fbs', 0))
        fbo = int(stock_data.get('fbo', 0))
    except Exception:
        fbs = 0
        fbo = 0
    return _build_price_row(product, price, fbs + fbo, fbs, fbo)


@router.post('/bulk-update', response_model=PriceListOut)
def bulk_update_prices(
    payload: PricesBulkUpdateIn,
    store_id: int = Query(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    owned_store_ids = _get_owned_ozon_store_ids(db, current_user.id)
    if store_id not in owned_store_ids:
        raise HTTPException(status_code=404, detail='Store not found')

    if not payload.updates:
        return PriceListOut(total=0, page=1, page_size=0, items=[], logs=_fetch_logs(db, current_user.id))

    updates_by_offer = {
        u.offer_id: {
            'new_price': round(u.new_price, 2),
            'min_price': round(u.min_price, 2) if u.min_price is not None else None,
        }
        for u in payload.updates
    }
    rows = db.execute(
        select(Product, Price)
        .join(Price, Price.product_id == Product.id)
        .where(
            Product.store_id == store_id,
            or_(Product.article.in_(list(updates_by_offer.keys())), Product.sku.in_(list(updates_by_offer.keys()))),
        )
    ).all()

    if not rows:
        return PriceListOut(total=0, page=1, page_size=0, items=[], logs=_fetch_logs(db, current_user.id))

    credentials = get_store_credentials(db, store_id)
    client = OzonClient()

    prices_payload: list[dict] = []
    requested_offer_ids: list[str] = []
    for product, _price in rows:
        offer = product.article or product.sku
        if not offer:
            continue
        update_values = updates_by_offer.get(offer)
        if update_values is None:
            continue
        requested_offer_ids.append(offer)
        ozon_data = _price.ozon_data or {}
        price_payload = ozon_data.get('price') or {}
        target_price = update_values['new_price']
        target_min_price = update_values['min_price']
        if target_min_price is None:
            target_min_price = _to_float(_price.min_price, target_price)
        target_min_price = round(min(target_price, max(target_min_price, 1.0)), 2)
        prices_payload.append(
            {
                'offer_id': offer,
                'price': f'{target_price:.2f}',
                'min_price': f'{target_min_price:.2f}',
                'currency_code': _to_str(price_payload.get('currency_code'), 'RUB'),
                'vat': _to_str(price_payload.get('vat'), '0.22'),
            }
        )

    if not prices_payload:
        return PriceListOut(total=0, page=1, page_size=0, items=[], logs=_fetch_logs(db, current_user.id))

    try:
        chunk_size = 1000
        for idx in range(0, len(prices_payload), chunk_size):
            client.import_prices(credentials, prices_payload[idx : idx + chunk_size])
        refreshed_items = client.fetch_prices(credentials, requested_offer_ids)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail='Не удалось обновить цены в Ozon') from exc

    refreshed_by_offer: dict[str, dict] = {}
    for item in refreshed_items:
        offer = _to_str(item.get('offer_id'))
        if offer:
            refreshed_by_offer[offer] = item

    updated_items: list[PriceRowOut] = []
    for product, price in rows:
        offer = product.article or product.sku
        if not offer or offer not in updates_by_offer:
            continue

        fresh_item = refreshed_by_offer.get(offer)
        target_values = updates_by_offer[offer]
        if fresh_item:
            fresh_price_payload = fresh_item.get('price') or {}
            updated_price = _to_float(fresh_price_payload.get('marketing_seller_price'), target_values['new_price'])
            updated_min_price = _to_float(fresh_price_payload.get('min_price'), target_values['min_price'] or price.min_price)
            price.ozon_data = fresh_item
        else:
            updated_price = target_values['new_price']
            updated_min_price = target_values['min_price'] or price.min_price

        price.previous_price = price.current_price
        price.current_price = updated_price
        price.min_price = round(min(updated_price, max(_to_float(updated_min_price, updated_price), 1.0)), 2)
        price.updated_at = datetime.utcnow()
        updated_items.append(_build_price_row(product, price))

    _log_price_action(
        db,
        current_user.id,
        'bulk_update',
        {'store_id': store_id, 'items': len(updated_items), 'status': 'updated_in_ozon'},
    )

    db.commit()
    return PriceListOut(total=len(updated_items), page=1, page_size=len(updated_items), items=updated_items, logs=_fetch_logs(db, current_user.id))


@router.get('/export-xlsx')
def export_prices_xlsx(
    store_id: int = Query(...),
    search: str | None = Query(default=None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    owned_store_ids = _get_owned_ozon_store_ids(db, current_user.id)
    if store_id not in owned_store_ids:
        raise HTTPException(status_code=404, detail='Store not found')

    rows = db.execute(_prices_query(db, store_id, search)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'prices'
    ws.append([
        'Артикул',
        'FBS',
        'FBO',
        'Остаток',
        'Цена',
        'Эквайринг',
        'Доставка',
        'Логистика',
        'Первая миля',
        'Упаковка',
        'Продвижение',
        'Комиссия %',
        'Комиссия руб',
        'Себестоимость',
        'Затраты на FBS',
        'К выплате',
        'Наценка',
        'Маржа',
        'Маржинальность',
    ])

    for product, price in rows:
        r = _build_price_row(product, price)
        ws.append([
            r.offer_id,
            r.fbs,
            r.fbo,
            r.stock,
            r.current_price,
            r.acquiring,
            r.customer_delivery,
            r.logistics,
            r.first_mile,
            r.packaging,
            r.promotion,
            r.ozon_commission_percent,
            r.ozon_commission_rub,
            r.cost_price,
            r.fbs_cost,
            r.payout_to_seller,
            r.markup_percent,
            r.margin_rub,
            r.margin_percent,
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    _log_price_action(db, current_user.id, 'export_xlsx', {'store_id': store_id, 'rows': len(rows)})
    db.commit()

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=ozon-prices.xlsx'},
    )


@router.post('/apply-markup', response_model=PriceListOut)
def apply_markup(
    payload: PricesApplyMarkupIn,
    store_id: int = Query(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if payload.markup_percent < payload.min_price_markup_percent:
        raise HTTPException(status_code=400, detail='markup_percent must be >= min_price_markup_percent')

    owned_store_ids = _get_owned_ozon_store_ids(db, current_user.id)
    if store_id not in owned_store_ids:
        raise HTTPException(status_code=404, detail='Store not found')

    q = select(Product, Price).join(Price, Price.product_id == Product.id).where(Product.store_id == store_id)
    if payload.offer_ids:
        q = q.where(or_(Product.article.in_(payload.offer_ids), Product.sku.in_(payload.offer_ids)))

    rows = db.execute(q).all()
    for _product, price in rows:
        ozon_data = price.ozon_data or {}
        source_price = ozon_data.get('price') or {}
        base_cost = _to_float(price.cost_price)
        if base_cost <= 0:
            base_cost = _to_float(source_price.get('net_price'), float(price.previous_price) if price.previous_price else float(price.current_price) * 0.7)
        commissions = ozon_data.get('commissions') or {}
        new_price = _calculate_price_from_markup(
            payload.markup_percent,
            base_cost,
            _to_float(commissions.get('fbs_deliv_to_customer_amount'), DEFAULT_CUSTOMER_DELIVERY),
            _to_float(commissions.get('fbs_direct_flow_trans_min_amount'), DEFAULT_LOGISTICS),
            _to_float(ozon_data.get('first_mile'), DEFAULT_FIRST_MILE),
            _to_float(ozon_data.get('packaging'), PACKAGING_COST),
            _to_float(ozon_data.get('promotion_percent'), PROMOTION_RATE * 100),
            _to_float(commissions.get('sales_percent_fbs'), DEFAULT_COMMISSION_PERCENT),
            _to_float(ozon_data.get('fbs_cost'), DEFAULT_FBS_COST),
        )
        min_price = _calculate_price_from_markup(
            payload.min_price_markup_percent,
            base_cost,
            _to_float(commissions.get('fbs_deliv_to_customer_amount'), DEFAULT_CUSTOMER_DELIVERY),
            _to_float(commissions.get('fbs_direct_flow_trans_min_amount'), DEFAULT_LOGISTICS),
            _to_float(ozon_data.get('first_mile'), DEFAULT_FIRST_MILE),
            _to_float(ozon_data.get('packaging'), PACKAGING_COST),
            _to_float(ozon_data.get('promotion_percent'), PROMOTION_RATE * 100),
            _to_float(commissions.get('sales_percent_fbs'), DEFAULT_COMMISSION_PERCENT),
            _to_float(ozon_data.get('fbs_cost'), DEFAULT_FBS_COST),
        )
        if new_price <= 0 or min_price <= 0:
            continue
        price.previous_price = price.current_price
        price.current_price = round(max(new_price, 1.0), 2)
        price.min_price = round(min(price.current_price, max(min_price, 1.0)), 2)
        price.updated_at = datetime.utcnow()

    _log_price_action(db, current_user.id, 'apply_markup', {'store_id': store_id, 'items': len(rows)})
    db.commit()

    refreshed = db.execute(q).all()
    items = [_build_price_row(product, price) for product, price in refreshed]
    return PriceListOut(total=len(items), page=1, page_size=len(items), items=items, logs=_fetch_logs(db, current_user.id))


@router.post('/import-cost-xlsx', response_model=PricesReloadOut)
def import_cost_price_xlsx(
    store_id: int = Query(...),
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    owned_store_ids = _get_owned_ozon_store_ids(db, current_user.id)
    if store_id not in owned_store_ids:
        raise HTTPException(status_code=404, detail='Store not found')

    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='Нужен файл .xlsx')

    try:
        wb = load_workbook(filename=BytesIO(file.file.read()), data_only=True)
        ws = wb.active
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail='Не удалось прочитать XLSX') from exc

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        raise HTTPException(status_code=400, detail='Файл пустой')

    header_map = {str(col).strip().lower(): idx for idx, col in enumerate(headers) if col is not None}
    offer_idx = header_map.get('offer_id')
    cost_idx = header_map.get('cost_price')
    if offer_idx is None or cost_idx is None:
        raise HTTPException(status_code=400, detail='В XLSX нужны колонки offer_id и cost_price')

    updates: dict[str, float] = {}
    for values in rows_iter:
        if not values:
            continue
        offer = _to_str(values[offer_idx] if offer_idx < len(values) else None)
        if not offer:
            continue
        cost = _to_float(values[cost_idx] if cost_idx < len(values) else None, default=-1)
        if cost < 0:
            continue
        updates[offer] = round(cost, 2)

    if not updates:
        raise HTTPException(status_code=400, detail='Нет валидных строк для обновления')

    matched_rows = db.execute(
        select(Product, Price)
        .join(Price, Price.product_id == Product.id)
        .where(
            Product.store_id == store_id,
            or_(Product.article.in_(list(updates.keys())), Product.sku.in_(list(updates.keys()))),
        )
    ).all()

    updated_count = 0
    for product, price in matched_rows:
        offer = product.article or product.sku
        if not offer or offer not in updates:
            continue
        price.cost_price = updates[offer]
        price.updated_at = datetime.utcnow()
        updated_count += 1

    _log_price_action(
        db,
        current_user.id,
        'import_cost_xlsx',
        {'store_id': store_id, 'rows_in_file': len(updates), 'updated_rows': updated_count},
    )
    db.commit()

    return PricesReloadOut(status='ok', message=f'Обновлена себестоимость для {updated_count} товаров')
